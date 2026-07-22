import io
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import app
from api.telegram_auth import require_telegram_auth
from generate_draft import save_project_record

TEST_USER_ID = "xlsx_export_test_user"
TEST_PROJECT = "xlsx_export_테스트현장"


def _fake_auth():
    return {"user_id": TEST_USER_ID, "username": "tester", "first_name": "테스트"}


def run():
    app.dependency_overrides[require_telegram_auth] = _fake_auth
    client = TestClient(app)
    results = []

    record = save_project_record(
        TEST_USER_ID, TEST_PROJECT, "위험성평가표", "테스트용 작업 정보",
        "| 항목 | 내용 |\n|------|------|\n| **현장명** | 테스트현장 |\n",
    )

    res = client.get(f"/projects/{TEST_PROJECT}/records/{record['id']}/export.xlsx")
    results.append(("엑셀 export 엔드포인트 200 응답", res.status_code == 200))
    results.append((
        "Content-Type이 xlsx MIME으로 설정됨",
        res.headers.get("content-type", "").startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    ))

    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    results.append((
        "응답 바이너리를 열어 표 내용 확인",
        ws.cell(row=1, column=1).value == "항목" and ws.cell(row=2, column=1).value == "현장명",
    ))

    res_404 = client.get(f"/projects/{TEST_PROJECT}/records/존재하지않는id/export.xlsx")
    results.append(("존재하지 않는 record_id는 404", res_404.status_code == 404))

    app.dependency_overrides.clear()

    all_ok = True
    for name, ok in results:
        print(f"[{'PASS' if ok else 'FAIL'}] {name}")
        all_ok = all_ok and ok
    print()
    print("전체 결과:", "PASS" if all_ok else "FAIL (위 로그 확인)")
    return all_ok


if __name__ == "__main__":
    run()
