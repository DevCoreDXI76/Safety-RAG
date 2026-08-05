"""
저장된 프로젝트 기록(record["draft"]의 Markdown 텍스트)을 파싱해
openpyxl 워크북으로 바인딩하고, xlsx 파일 바이트를 반환한다.

셀 스타일(열너비/줄바꿈/테두리/조건부서식/인쇄설정)은 docs/샘플문서/의 5개
서식목업(위험성평가표·표준 작업계획서·TBM 일지·안전보건교육일지·
산업안전보건관리비 사용명세서)을 실측해 그대로 반영했다.
"""

import io
import math
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from document_styles import (
    AI_SCORE_FOOTNOTE, DEFAULT_COLUMN_WIDTH, PORTRAIT_DOCUMENT_TYPES, base_header, get_style,
    parse_ai_score_cell, CENTER_ALIGN_HEADERS,
)
from markdown_tables import parse_markdown_blocks

_TITLE_FONT = Font(size=28, bold=True, underline="single")
_ALIGN_TITLE_DOC = Alignment(horizontal="center", vertical="center")

_BOX_TITLE_FONT = Font(size=16, bold=True, color="000000")
# wrap_text 누락 시 긴 박스 제목(예: "7. 절연용 보호구 및 방호구 등 준비·점검·
# 착용·사용에 관한 사항")이 줄바꿈 없이 한 줄로 나가면서 인쇄 페이지 폭을
# 넘겨 끝부분이 잘려 보인다(2026-08-05 4차 피드백). 행 높이는 이미
# _set_row_height로 이 제목 길이만큼 계산해두므로, wrap_text만 켜면 그 안에서
# 자연스럽게 줄바꿈된다.
_ALIGN_TITLE = Alignment(horizontal="left", vertical="center", wrap_text=True)

_BODY_FONT_SIZE = 12
_HEADER_FONT = Font(size=_BODY_FONT_SIZE, bold=True)
_BODY_FONT = Font(size=_BODY_FONT_SIZE)
_FOOTNOTE_FONT = Font(size=9, color="808080")
_ALIGN_FOOTNOTE = Alignment(horizontal="left", vertical="center")

_THIN_SIDE = Side(style="thin")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

_INVALID_SHEET_CHARS_RE = re.compile(r"[:\\/?*\[\]]")
_COMMENT_AUTHOR = "safety-rag"

# 모든 내용을 B열부터 쓰고 A열은 공란(여백)으로 남긴다(2026-08-05 요청).
_COL_OFFSET = 1
_COL_A_WIDTH = 3

_CONTENT_AWARE_WIDTH_DOCUMENT_TYPES = {"위험성평가표"}
_MIN_EXCEL_COL_WIDTH = 6
_MAX_EXCEL_COL_WIDTH = 60
_EXCEL_COL_WIDTH_PADDING = 2
_LINE_BREAK_RE = re.compile(r"<br\s*/?>|\n")


def _text_width_units(text):
    """한글 등 전각 문자는 2칸, 그 외(영문·숫자 등)는 1칸으로 계산한 근사 폭."""
    lines = _LINE_BREAK_RE.split(text) if text else [""]
    widest = 0
    for line in lines:
        units = sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in line)
        widest = max(widest, units)
    return widest


def _content_aware_excel_widths(table_blocks, ncols):
    """
    표 전체(같은 열 위치를 공유하는 모든 표)에서 각 열의 가장 넓은 셀 내용을
    기준으로 엑셀 열 폭을 정한다(2026-08-05 요청, 위험성평가표 전용).

    폭 측정은 실제로 셀에 표시되는 값 기준이어야 한다 — 빈도·강도·위험등급처럼
    한 글자/한 단어만 들어가는 열도 원본 마크다운 셀 텍스트에는 "B(AI 제안값,
    현장 확인 필수)"처럼 안내문이 붙어 있어(parse_ai_score_cell로 나중에
    분리·제거됨), 그 원문 그대로 폭을 재면 실제로는 좁아야 할 열이 크게
    부풀려진다(2026-08-05 4차 피드백 "빈도·강도·위험등급 등의 셀 가로크기를
    셀 제목의 크기만큼 줄일 것" — 원인이 바로 이것이었음).
    """
    widths = [_MIN_EXCEL_COL_WIDTH] * ncols
    for tb in table_blocks:
        for row in tb["rows"]:
            for col in range(min(len(row), ncols)):
                number, _note = parse_ai_score_cell(row[col])
                measure_text = str(number) if number is not None else row[col]
                widths[col] = max(widths[col], _text_width_units(measure_text))
    return [min(w + _EXCEL_COL_WIDTH_PADDING, _MAX_EXCEL_COL_WIDTH) for w in widths]


# 엑셀은 "병합된" 셀에서는 wrap_text가 켜져 있어도 행 높이를 자동으로 늘려주지
# 않는다(잘 알려진 엑셀 제약) — 그래서 긴 문단이 한 줄로 눌려 보이거나, 심하면
# 스크롤하다 "내용이 통째로 빠졌다"고 오해하기 쉽다(2026-08-05 실사용 피드백).
# 행 높이를 직접 계산해서 지정한다.
_EXCEL_LINE_HEIGHT_PT = 16
_EXCEL_ROW_HEIGHT_PADDING_PT = 6

# 엑셀 "열너비" 단위는 통합문서 기본폰트(Calibri 11pt) 기준으로 정의되는데,
# 본문 셀은 그보다 큰 12pt인 데다 Calibri에는 한글 글리프가 없어 실제로는
# 맑은 고딕 등으로 자동 치환되어 렌더링 폭이 더 넓다. 이 차이를 그대로
# 두면 줄바꿈 줄 수를 실제보다 적게 추정해 텍스트가 셀 아래로 잘려 보인다
# (2026-08-05 2차 피드백 "셀 크기 때문에 텍스트가 잘려서 안 보임"). 안전마진을
# 곱해 조금 더 넉넉하게(줄이 더 필요하다고) 계산한다.
_WIDTH_ESTIMATE_SAFETY_FACTOR = 1.25


def _wrapped_line_count(text, span_width_units, font_size=_BODY_FONT_SIZE):
    """
    span_width_units 폭에 줄바꿈해서 넣을 때 필요한 줄 수(최소 1, 안전마진
    적용). font_size가 본문(12pt)보다 크면(예: 박스 제목 16pt) 같은 열 폭에
    실제로 들어가는 글자 수가 더 적으므로, 폰트 크기 비율만큼 폭 추정치를
    더 넓게 잡는다.
    """
    if not text:
        return 1
    font_scale = font_size / _BODY_FONT_SIZE
    total = 0
    for line in _LINE_BREAK_RE.split(str(text)):
        width = _text_width_units(line) * _WIDTH_ESTIMATE_SAFETY_FACTOR * font_scale
        total += max(1, math.ceil(width / max(span_width_units, 1)))
    return max(total, 1)


def _set_row_height(ws, row, cell_texts_and_widths, font_size=_BODY_FONT_SIZE):
    """
    그 행의 셀들 중 가장 많은 줄 수가 필요한 셀 기준으로 행 높이를 지정한다.
    font_size가 본문보다 크면(박스 제목 16pt) 줄 높이도 그 비율만큼 키운다 —
    안 그러면 2줄로 줄바꿈된 제목의 위쪽이 바로 위 행에 눌려 잘려 보인다
    (2026-08-05 5차 피드백 "제목의 텍스트 윗 부분이 위쪽 셀에 의해서 잘림").
    """
    max_lines = max(
        (_wrapped_line_count(text, width, font_size) for text, width in cell_texts_and_widths), default=1
    )
    line_height_pt = _EXCEL_LINE_HEIGHT_PT * (font_size / _BODY_FONT_SIZE)
    ws.row_dimensions[row].height = max_lines * line_height_pt + _EXCEL_ROW_HEIGHT_PADDING_PT


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


# "(빈칸 - 현장 기재)"류 안내문은 실제 데이터가 아니라 현장에서 채워야 할
# 자리표시자이므로, PDF(export_pdf.py의 _PLACEHOLDER_RE)와 동일하게 연한
# 회색으로 구분해 표시한다(2026-08-05 3차 피드백 "모든 엑셀문서에 동일하게
# 적용해줘").
_PLACEHOLDER_RE = re.compile(r"\([^()]*빈칸[^()]*\)")
_PLACEHOLDER_COLOR = "999999"


def _placeholder_rich_text(text, font):
    """
    text 안의 "(빈칸 - 현장 기재)"류 문구만 연한 회색으로 강조한 rich text를
    만든다. 나머지 구간은 font(이미 그 셀에 적용된 폰트)를 그대로 유지한다.
    """
    base_color = font.color.rgb if font.color else "FF000000"
    normal_run = InlineFont(sz=font.size, b=font.bold, color=base_color)
    gray_run = InlineFont(sz=font.size, b=font.bold, color=_PLACEHOLDER_COLOR)
    blocks = []
    last_end = 0
    for m in _PLACEHOLDER_RE.finditer(text):
        if m.start() > last_end:
            blocks.append(TextBlock(normal_run, text[last_end:m.start()]))
        blocks.append(TextBlock(gray_run, m.group(0)))
        last_end = m.end()
    if last_end < len(text):
        blocks.append(TextBlock(normal_run, text[last_end:]))
    return CellRichText(*blocks)


def _sheet_title(document_type):
    """엑셀 시트명 제약(31자 이내, : \\ / ? * [ ] 금지)을 만족하도록 정리한다."""
    cleaned = _INVALID_SHEET_CHARS_RE.sub("", document_type)
    return cleaned[:31] or "문서"


# ECMA-376 표준 용지 코드(paperSize) — 9 = A4. openpyxl에는 이름 상수가
# 없어 숫자 코드를 직접 쓴다. 값을 안 주면 뷰어/프린터 환경에 따라 A4가
# 아닌 용지로 인쇄될 수 있어(2026-08-05 "A4에 안 맞게 나옴" 피드백) 항상 명시한다.
_PAPER_SIZE_A4 = 9

# 위험성평가표는 표가 여러 개라 스크롤이 잦은데, 틀고정이 오히려 헤더 열을
# 가려 혼란을 준다는 피드백(2026-08-05)으로 이 문서유형만 틀고정을 끈다.
_NO_FREEZE_DOCUMENT_TYPES = {"위험성평가표"}
# 표준 작업계획서는 문서가 길어 스크롤하면 "작업 개요" 박스가 안 보인다는
# 피드백(2026-08-05) — 그 박스(첫 번째 표) 바로 다음 행에 틀고정한다.
_FREEZE_AFTER_FIRST_TABLE_DOCUMENT_TYPES = {"표준 작업계획서"}

# "인쇄 미리보기 여백이 너무 많다"(2026-08-05 2차 피드백) — 상하좌우 25px
# (96dpi 기준 화면 픽셀 환산, 엑셀 인쇄여백 단위인 인치로 변환)로 축소한다.
# 머리글/바닥글 텍스트는 실제로 쓰지 않으므로 그 여백도 함께 줄인다.
_PRINT_MARGIN_IN = 25 / 96
_PRINT_HEADER_FOOTER_MARGIN_IN = 0.2

# 인쇄 배율 계산에 쓰는 A4 가로폭(인치)과 열너비 단위→픽셀 환산 근사치.
# openpyxl은 이 변환을 제공하지 않아 통합문서 기본폰트(Calibri 11pt) 기준의
# 관용적 근사식(문자당 7px + 셀 여백/테두리 5px)을 그대로 쓴다 — 정확한 1:1
# 변환은 아니지만 "얼마나 남는지/넘치는지" 판단에는 충분하다.
_A4_WIDTH_IN = {"portrait": 8.27, "landscape": 11.69}
_EXCEL_COL_WIDTH_PX_SLOPE = 7
_EXCEL_COL_WIDTH_PX_INTERCEPT = 5
# 확대 배율 상한 — 이전에는 115였는데, 정적 스펙 열이 4개(TBM 일지)인
# 문서에서 115%로도 실제 인쇄에서 우측 열이 다음 페이지로 밀려나는(페이지가
# 6장으로 늘어남) 문제가 확인됐다(2026-08-05 6차 피드백). 열너비 단위→픽셀
# 근사식 자체의 오차가 생각보다 커서, 상한을 더 보수적으로 낮춘다.
_PRINT_SCALE_MAX_PERCENT = 105
# "페이지가 안 나눠지도록 A4 가로 사이즈에 다 들어오게 해달라"(2026-08-05
# 4차 피드백)는 요구가 최우선이라, 최소 배율에는 사실상 하한을 두지 않는다
# (음수/0 같은 비정상값만 막는 안전망 수준). 위험성평가표는 열이 최대 13개라
# 필요하면 이보다 훨씬 작게도 줄어들어야 한 페이지 폭에 들어간다.
_PRINT_SCALE_MIN_PERCENT = 15
# 인쇄 배율 계산 전용 안전마진 — 행높이 계산의 안전마진
# (_WIDTH_ESTIMATE_SAFETY_FACTOR, 1.25배)을 그대로 가져다 쓰면 실제로
# 필요한 것보다 지나치게 많이 줄어든다는 피드백을 받아(2026-08-05 5차
# "인쇄 배율을 50%까지 늘려줘") 한 차례 낮췄었다. 그런데 실제 위험성평가표
# 인쇄에서 열 1개(재평가 필요여부)가 다음 페이지로 밀려나는 게 확인되어
# (2026-08-05 6차 "인쇄배율만 조금 조정하면 될 것 같다") 다시 소폭 올린다.
_PRINT_SCALE_WIDTH_SAFETY_FACTOR = 1.0
# 반올림/렌더링 오차로 실제 폭이 계산값보다 살짝 더 넓어도 페이지를 넘기지
# 않도록, 계산값을 항상 내림(floor)한 뒤 몇 %p 더 줄여서 적용한다.
_PRINT_SCALE_SAFETY_BUFFER_PERCENT = 5

# 위험성평가표는 열이 많아(최대 13열) 폭이 페이지를 넘기기 쉽고, 그때 배율을
# 직접 계산해서 명시하지 않으면 한 페이지 폭에 들어오지 못하고 좌우로
# 페이지가 나뉘는 문제가 실사용 인쇄 미리보기에서 확인됐다(2026-08-05 4차
# 피드백) — 이 문서유형은 계산된 배율을 항상 그대로 적용한다(축소든 확대든).
_EXPLICIT_SCALE_DOCUMENT_TYPES = {"위험성평가표"}
# 표준 작업계획서는 정적 스펙 열너비가 좁아(대부분 항목/내용 2열짜리 kv표뿐)
# 우측에 여백이 많이 남는다는 피드백(2026-08-05 3차)이 있었고, 확대 후에도
# 별다른 문제 제기가 없어 이 문서유형만 "계산된 배율이 100%를 넘길 때만
# 확대" 모드를 유지한다.
# TBM 일지는 표가 4열(핵심 위험요인/참석자 명단)까지 있어, 같은 방식으로
# 확대했을 때 실제 인쇄에서 우측 열이 다음 페이지로 밀려나는 문제가 두
# 라운드 연속으로 재현됐다(2026-08-05 5·6차) — 근사식의 오차 폭 안에서
# 안전하게 확대할 배율을 특정하기 어렵다고 판단해, 확대를 포기하고 뷰어의
# "폭에 맞춤" 자동 축소(항상 안전했던 원래 방식)로 되돌린다.
_ENLARGE_ONLY_SCALE_DOCUMENT_TYPES = {"표준 작업계획서"}


def _print_scale_percent(document_type, column_widths, max_col_count):
    """
    문서유형이 실제 열 폭 기준으로 A4 인쇄가능 폭에 맞으려면 몇 %가 필요한지
    계산한다(축소 필요 시 100% 미만, 확대해도 되면 100% 초과). 호출부
    (_apply_print_settings)가 문서유형별로 이 값을 실제로 적용할지 결정한다.
    """
    total_units = _COL_A_WIDTH + sum(column_widths[:max_col_count])
    if total_units <= 0:
        return 100
    content_px = (
        total_units * _EXCEL_COL_WIDTH_PX_SLOPE + _EXCEL_COL_WIDTH_PX_INTERCEPT
    ) * _PRINT_SCALE_WIDTH_SAFETY_FACTOR
    orientation = "portrait" if document_type in PORTRAIT_DOCUMENT_TYPES else "landscape"
    usable_in = _A4_WIDTH_IN[orientation] - 2 * _PRINT_MARGIN_IN
    usable_px = usable_in * 96
    scale = math.floor(usable_px / content_px * 100) - _PRINT_SCALE_SAFETY_BUFFER_PERCENT
    return max(_PRINT_SCALE_MIN_PERCENT, min(scale, _PRINT_SCALE_MAX_PERCENT))


def _apply_print_settings(ws, document_type, scale_percent=100):
    """
    5개 서식목업 공통 인쇄설정: A4, 문서유형별 방향, 여백.
    - 위험성평가표: 계산된 배율을 항상 명시(_EXPLICIT_SCALE_DOCUMENT_TYPES)
    - 표준 작업계획서·TBM 일지: 계산된 배율이 100%를 넘길 때만(확대) 명시,
      아니면 뷰어의 "폭에 맞춤" 자동 축소(_ENLARGE_ONLY_SCALE_DOCUMENT_TYPES)
    - 그 외: 항상 뷰어의 "폭에 맞춤" 자동 축소

    print_title_rows(반복할 행)는 의도적으로 설정하지 않는다 — 이 시트는
    표 여러 개와 서술형 문단이 섞여 있어서, 그중 한 표의 헤더 행을 "반복할
    행"으로 지정하면 그 표의 데이터가 끝난 뒤 전혀 무관한 뒤쪽 섹션(예: TBM
    일지 "3. 중점 위험요인")이 시작되는 페이지에도 그 표 헤더가 엉뚱하게
    다시 찍혀 나온다(2026-08-05 7차 피드백 "중점위험 요인 다음에 잘못된
    박스가 들어감" — 실제로는 잘못된 셀이 아니라 이 인쇄 반복행 기능이
    원인이었음). 화면 스크롤용 틀고정(freeze_panes)은 이 문제와 무관해
    그대로 유지한다.
    """
    ws.page_setup.orientation = "portrait" if document_type in PORTRAIT_DOCUMENT_TYPES else "landscape"
    ws.page_setup.paperSize = _PAPER_SIZE_A4
    use_explicit_scale = (
        document_type in _EXPLICIT_SCALE_DOCUMENT_TYPES
        or (document_type in _ENLARGE_ONLY_SCALE_DOCUMENT_TYPES and scale_percent > 100)
    )
    if use_explicit_scale:
        ws.sheet_properties.pageSetUpPr.fitToPage = False
        ws.page_setup.scale = scale_percent
    else:
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = _PRINT_MARGIN_IN
    ws.page_margins.right = _PRINT_MARGIN_IN
    ws.page_margins.top = _PRINT_MARGIN_IN
    ws.page_margins.bottom = _PRINT_MARGIN_IN
    ws.page_margins.header = _PRINT_HEADER_FOOTER_MARGIN_IN
    ws.page_margins.footer = _PRINT_HEADER_FOOTER_MARGIN_IN


def record_to_xlsx_bytes(record):
    """
    record["draft"]에서 헤딩(박스 제목)·표·서술형 문단을 순서대로 시트에
    채운다. 맨 위에 문서 제목(PDF와 동일하게 28pt·굵게·밑줄)을 두고, 모든
    내용은 A열을 공란으로 남긴 채 B열부터 쓴다(2026-08-05 요청). 표가
    여러 개면 표 사이에 빈 행 하나를 둔다. draft가 완전히 비어있는
    경우(이론상 거의 없음)만 원문 그대로 A1에 넣는다.
    """
    blocks = parse_markdown_blocks(record["draft"])
    document_type = record["document_type"]
    style = get_style(document_type)

    header_fill = _fill(style.header_fill)
    data_header_font = Font(size=_BODY_FONT_SIZE, bold=True, color=style.header_font_color)
    kv_header_fill = _fill(style.kv_header_fill)
    risk_fills = {grade: _fill(hex_color) for grade, hex_color in style.risk_grade_colors.items()}

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(document_type)

    if not blocks:
        ws.cell(row=1, column=1, value=record["draft"])
        _apply_print_settings(ws, document_type)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # 박스 가로 크기를 모든 표가 동일하게 쓰도록, 시트 전체에서 가장 열이 많은
    # 표 기준으로 폭을 맞춘다(2026-08-05 요청) — kv표(항목/내용, 2열)의 "내용"
    # 칸은 이 값까지 병합해서 다른 표와 박스 너비가 맞아 보이게 한다.
    table_blocks = [b for b in blocks if b["type"] == "table"]
    max_col_count = max((len(row) for tb in table_blocks for row in tb["rows"]), default=1)

    if document_type in _CONTENT_AWARE_WIDTH_DOCUMENT_TYPES:
        column_widths = _content_aware_excel_widths(table_blocks, max_col_count)
    else:
        column_widths = list(style.column_widths)
        while len(column_widths) < max_col_count:
            column_widths.append(DEFAULT_COLUMN_WIDTH)

    ws.column_dimensions["A"].width = _COL_A_WIDTH

    # "표준 작업계획서 (전기작업)"처럼 세부 작업유형을 제목 옆에 표기한다
    # (2026-08-05 요청). work_type이 없는 기존 기록/다른 문서유형은 그대로 둔다.
    work_type = record.get("work_type")
    title_text = f"{document_type} ({work_type})" if work_type else document_type

    title_cell = ws.cell(row=1, column=1 + _COL_OFFSET, value=title_text)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _ALIGN_TITLE_DOC
    ws.merge_cells(
        start_row=1, start_column=1 + _COL_OFFSET, end_row=1, end_column=max_col_count + _COL_OFFSET
    )

    current_row = 2
    table_index = 0
    freeze_row = None
    first_table_next_row = None  # 첫 번째 표(+각주) 바로 다음 행 — 작업계획서 틀고정 기준
    risk_score_ranges = []  # (col_letter, first_data_row, last_data_row)

    for block in blocks:
        if block["type"] == "heading":
            title_cell = ws.cell(row=current_row, column=1 + _COL_OFFSET, value=block["text"])
            title_cell.font = _BOX_TITLE_FONT
            title_cell.alignment = _ALIGN_TITLE
            if _PLACEHOLDER_RE.search(block["text"]):
                title_cell.value = _placeholder_rich_text(block["text"], title_cell.font)
            if max_col_count > 1:
                ws.merge_cells(
                    start_row=current_row, start_column=1 + _COL_OFFSET,
                    end_row=current_row, end_column=max_col_count + _COL_OFFSET,
                )
            _set_row_height(
                ws, current_row, [(block["text"], sum(column_widths[:max_col_count]))],
                font_size=_BOX_TITLE_FONT.size,
            )
            current_row += 1
            continue

        if block["type"] == "text":
            text_cell = ws.cell(row=current_row, column=1 + _COL_OFFSET, value=block["text"])
            text_cell.font = _BODY_FONT
            text_cell.alignment = _ALIGN_LEFT_TOP
            if _PLACEHOLDER_RE.search(block["text"]):
                text_cell.value = _placeholder_rich_text(block["text"], text_cell.font)
            if max_col_count > 1:
                ws.merge_cells(
                    start_row=current_row, start_column=1 + _COL_OFFSET,
                    end_row=current_row, end_column=max_col_count + _COL_OFFSET,
                )
            _set_row_height(ws, current_row, [(block["text"], sum(column_widths[:max_col_count]))])
            current_row += 2  # 다음 블록과 구분되도록 빈 행 하나
            continue

        table = block["rows"]
        header_row = current_row
        headers_base = [base_header(h) for h in table[0]]
        is_kv_table = len(table[0]) == 2
        risk_grade_col_idxs = []
        if not is_kv_table:
            for idx, header_text in enumerate(headers_base, start=1):
                if header_text in style.risk_grade_headers:
                    risk_grade_col_idxs.append(idx)

        ai_value_present = False

        for row_offset, row_cells in enumerate(table):
            is_header = row_offset == 0
            row_texts_and_widths = []
            for col_idx, value in enumerate(row_cells, start=1):
                number, note = parse_ai_score_cell(value)
                display_value = number if number is not None else value
                cell = ws.cell(row=current_row, column=col_idx + _COL_OFFSET, value=display_value)
                if note:
                    cell.comment = Comment(note, _COMMENT_AUTHOR)
                    ai_value_present = True
                cell.border = _THIN_BORDER

                if is_kv_table and col_idx >= 2:
                    span_width = sum(column_widths[1:max_col_count])
                else:
                    span_width = column_widths[col_idx - 1] if col_idx - 1 < len(column_widths) else DEFAULT_COLUMN_WIDTH
                row_texts_and_widths.append((str(display_value) if display_value is not None else "", span_width))

                if is_kv_table:
                    if col_idx == 1:
                        cell.font = _HEADER_FONT
                        cell.fill = kv_header_fill
                        cell.alignment = _ALIGN_CENTER
                    else:
                        if is_header:
                            cell.font = _HEADER_FONT
                            cell.fill = kv_header_fill
                        else:
                            cell.font = _BODY_FONT
                        cell.alignment = _ALIGN_LEFT_CENTER
                elif is_header:
                    cell.font = data_header_font
                    cell.fill = header_fill
                    cell.alignment = _ALIGN_CENTER
                else:
                    cell.font = _BODY_FONT
                    header_text = headers_base[col_idx - 1] if col_idx - 1 < len(headers_base) else ""
                    cell.alignment = (
                        _ALIGN_CENTER if header_text.lower() in CENTER_ALIGN_HEADERS else _ALIGN_LEFT_TOP
                    )
                    if col_idx not in risk_grade_col_idxs:
                        # "위험성 추정 행렬" 참고표처럼 헤더명이 "위험등급"이 아니라
                        # risk_grade_column_indices로는 안 잡히지만, 셀 값 자체가
                        # 등급 문자(A/B/C)인 표도 본문과 같은 색으로 칠한다
                        # (2026-08-05 요청).
                        grade = str(display_value).strip().upper() if display_value is not None else ""
                        if grade in risk_fills:
                            cell.fill = risk_fills[grade]
                            cell.alignment = _ALIGN_CENTER

                if isinstance(display_value, str) and _PLACEHOLDER_RE.search(display_value):
                    cell.value = _placeholder_rich_text(display_value, cell.font)

            if is_kv_table and max_col_count > 2:
                ws.merge_cells(
                    start_row=current_row, start_column=2 + _COL_OFFSET,
                    end_row=current_row, end_column=max_col_count + _COL_OFFSET,
                )

            _set_row_height(ws, current_row, row_texts_and_widths)
            current_row += 1

            if table_index == 1 and is_header:
                freeze_row = current_row

        if risk_grade_col_idxs and len(table) > 1:
            for idx in risk_grade_col_idxs:
                risk_score_ranges.append(
                    (get_column_letter(idx + _COL_OFFSET), header_row + 1, current_row - 1)
                )

        # AI 제안값이 있었던 표는 PDF/HWPX와 동일하게 각주를 표 바로 아래에 남긴다
        # (2026-08-05 요청 — XLSX만 이 각주가 빠져 있었음).
        if ai_value_present:
            footnote_cell = ws.cell(row=current_row, column=1 + _COL_OFFSET, value=AI_SCORE_FOOTNOTE)
            footnote_cell.font = _FOOTNOTE_FONT
            footnote_cell.alignment = _ALIGN_FOOTNOTE
            if max_col_count > 1:
                ws.merge_cells(
                    start_row=current_row, start_column=1 + _COL_OFFSET,
                    end_row=current_row, end_column=max_col_count + _COL_OFFSET,
                )
            current_row += 1

        if table_index == 0:
            first_table_next_row = current_row + 1

        table_index += 1
        current_row += 1  # 표 사이 빈 행

    if document_type in _NO_FREEZE_DOCUMENT_TYPES:
        freeze_row = None
    elif document_type in _FREEZE_AFTER_FIRST_TABLE_DOCUMENT_TYPES:
        freeze_row = first_table_next_row or 2
    elif freeze_row is None:
        freeze_row = 2

    for col_idx in range(1, max_col_count + 1):
        ws.column_dimensions[get_column_letter(col_idx + _COL_OFFSET)].width = column_widths[col_idx - 1]

    if freeze_row is not None:
        ws.freeze_panes = f"A{freeze_row}"

    for col_letter, first_row, last_row in risk_score_ranges:
        cell_range = f"{col_letter}{first_row}:{col_letter}{last_row}"
        for grade, fill in risk_fills.items():
            ws.conditional_formatting.add(
                cell_range, CellIsRule(operator="equal", formula=[f'"{grade}"'], fill=fill)
            )

    scale_percent = _print_scale_percent(document_type, column_widths, max_col_count)
    _apply_print_settings(ws, document_type, scale_percent=scale_percent)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
