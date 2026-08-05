import io
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
from document_styles import AI_SCORE_FOOTNOTE, STYLE_SPECS
from export_xlsx import _print_scale_percent, record_to_xlsx_bytes


def _placeholder_run_present(rich_value):
    """rich_value(CellRichText) 안에 '빈칸'을 포함하면서 999999(연회색)로
    칠해진 run이 있는지 확인한다."""
    if not isinstance(rich_value, CellRichText):
        return False
    for block in rich_value:
        rgb = getattr(getattr(block.font, "color", None), "rgb", "") or ""
        if "빈칸" in block.text and rgb.upper().endswith("999999"):
            return True
    return False

SAMPLE_RECORD = {
    "id": "abc123",
    "document_type": "위험성평가표",
    "project_info": "강남지사 광케이블 지중매설",
    "draft": (
        "# 위험성평가표 초안\n\n"
        "## ■ 기본 정보\n\n"
        "| 항목 | 내용 |\n"
        "|------|------|\n"
        "| **현장명** | 강남지사_광케이블 |\n"
        "| **작성일** | 2026-07-20 |\n\n"
        "## ■ 위험요인\n\n"
        "| 순번 | 위험요인 | 위험성 |\n"
        "|------|----------|--------|\n"
        "| 1 | 추락 | 12 (AI 제안값, 현장 확인 필수) |\n"
        "| 2 | 감전 | 9 (AI 제안값, 현장 확인 필수) |\n"
    ),
    "created_at": "2026-07-22 10:00:00",
}

# 2026-08-05 실사용 XLSX 피드백: 박스 제목(헤딩)이 통째로 사라지고, 표가 아닌
# 서술형 섹션도 사라지며, kv표(항목/내용)와 다중열표의 박스 가로 크기가
# 서로 달라 보이는 문제.
SAMPLE_RECORD_WITH_HEADING_AND_PROSE = {
    "id": "heading1",
    "document_type": "TBM 일지",
    "project_info": "박스 제목/서술형 섹션 검증용",
    "draft": (
        "# TBM 일지 초안\n\n"
        "## ■ 기본 정보\n\n"
        "| 항목 | 내용 |\n"
        "|------|------|\n"
        "| 현장명 | 강남지사 |\n\n"
        "## ■ 핵심 위험요인\n\n"
        "| 번호 | 유해위험요인 | 대책 |\n"
        "|------|------|------|\n"
        "| 1 | 감전 | 절연장갑 착용 |\n\n"
        "### 3. 중점(One Point) 위험요인\n\n"
        "오늘은 활선 근접 작업이 포함되어 있으므로 무전압 상태를 반드시 확인한다.\n"
    ),
    "created_at": "2026-08-05 10:00:00",
}

SAMPLE_RECORD_NO_TABLE = {
    "id": "def456",
    "document_type": "기타",
    "project_info": "표 없는 문서",
    "draft": "이 문서에는 표가 없습니다.",
    "created_at": "2026-07-22 10:00:00",
}

SAMPLE_RECORD_WITH_SCORE = {
    "id": "score1",
    "document_type": "위험성평가표",
    "project_info": "테스트용 점수 셀",
    "draft": (
        "| 위험요인 | 빈도 | 강도 | 위험등급 | 개선후 위험등급 |\n"
        "|----------|------|------|----------|------------------|\n"
        "| 지게차 충돌 | 3(AI 제안값, 현장 확인 필수) | 2(AI 제안값, 현장 확인 필수) | "
        "A(AI 제안값, 현장 확인 필수) | B(AI 제안값, 현장 확인 필수) |\n"
    ),
    "created_at": "2026-07-22 10:00:00",
}

# 2026-08-05 요청: 위험성평가표는 열 폭을 셀 내용 글자수 기반으로 조정.
# "빈도" 열은 짧은 숫자만, "위험성 감소대책"은 긴 서술형 문장 — 폭이 크게
# 달라야 한다.
SAMPLE_RECORD_RISK_WIDE_NARROW = {
    "id": "widthtest1",
    "document_type": "위험성평가표",
    "project_info": "열 폭 자동조절 검증용",
    "draft": (
        "| 단위작업 | 빈도 | 위험성 감소대책 상세 설명 |\n"
        "|------|------|------|\n"
        "| 굴착 | 3 | 흙막이 지보공 설치, 구배 기준 준수, 굴착 깊이 확인 등 상세한 대책을 서술한다 |\n"
    ),
    "created_at": "2026-08-05 10:00:00",
}

# 2026-08-05 요청: 엑셀은 병합된 셀에서 줄바꿈(wrap_text)이 있어도 행 높이를
# 자동으로 늘려주지 않는다 — 그래서 종합의견 등 긴 문단이 한 줄로 눌려 보이고,
# 심하면 "내용이 통째로 빠졌다"고 오해하게 만든다. 행 높이를 직접 계산해서
# 지정해야 한다.
SAMPLE_RECORD_LONG_TEXT = {
    "id": "longtext1",
    "document_type": "위험성평가표",
    "project_info": "행 높이 자동조절 검증용",
    "draft": (
        "## ■ 종합의견\n\n"
        "본 작업은 감전 및 아크 관련 위험요인이 가장 중대하다. 전원 차단·무전압 확인·잠금장치 적용 등 "
        "전기작업 안전절차를 최우선으로 준수해야 하며, 작업 전 전기안전작업계획서 및 회로도 확인을 "
        "반드시 병행할 것을 권고한다. 중량물 하역 및 개구부 주변 작업 시에도 협착·추락 방지대책을 "
        "함께 이행해야 한다.\n\n"
        "## ■ 기본 정보\n\n"
        "| 항목 | 내용 |\n"
        "|------|------|\n"
        "| 짧은값 | 예 |\n"
        "| 긴값 | 매설물 관리기관 확인 및 이설·보호대책 수립, 굴착 착수 전 관계 기관 협의 후 착공계 제출 |\n"
    ),
    "created_at": "2026-08-05 10:00:00",
}

# 2026-08-05 요청: "위험성 추정 행렬" 참고표는 열 헤더가 "위험등급"이 아니라
# risk_grade_column_indices로 안 잡혀서 A/B/C 셀에 색이 안 들어가고 있었다.
SAMPLE_RECORD_RISK_MATRIX = {
    "id": "matrix1",
    "document_type": "위험성평가표",
    "project_info": "위험성 추정 행렬 색상 검증용",
    "draft": (
        "| 빈도\\강도 | 1(낮음) | 2(보통) | 3(높음) |\n"
        "|------|------|------|------|\n"
        "| 1(낮음) | C | C | B |\n"
        "| 2(보통) | C | B | A |\n"
        "| 3(높음) | B | A | A |\n"
    ),
    "created_at": "2026-08-05 10:00:00",
}

SAMPLE_RECORD_OTHER_DOC_TYPE = {
    "id": "widthtest2",
    "document_type": "TBM 일지",
    "project_info": "다른 문서유형은 정적 스펙 유지 확인용",
    "draft": (
        "| 항목 | 내용 |\n"
        "|------|------|\n"
        "| 일자 | 2026-08-05 |\n"
    ),
    "created_at": "2026-08-05 10:00:00",
}

# 2026-08-05 요청: 위험성평가표는 틀고정 제거, 작업계획서는 "작업 개요"
# 박스(첫 표) 바로 다음에 틀고정. 두 박스 이상을 가진 draft로 검증한다.
SAMPLE_RECORD_FREEZE_RISK = {
    "id": "freeze1",
    "document_type": "위험성평가표",
    "project_info": "틀고정 제거 검증용",
    "draft": (
        "## ■ 기본 정보\n\n| 항목 | 내용 |\n|------|------|\n| 현장명 | 강남 |\n\n"
        "## ■ 위험요인\n\n| 순번 | 위험요인 | 위험성 |\n|------|------|------|\n| 1 | 감전 | 9 |\n"
    ),
    "created_at": "2026-08-05 10:00:00",
}

SAMPLE_RECORD_FREEZE_WORKPLAN = {
    "id": "freeze2",
    "document_type": "표준 작업계획서",
    "project_info": "작업 개요 박스 다음 틀고정 검증용",
    "draft": (
        "## ■ 작업 개요\n\n| 항목 | 내용 |\n|------|------|\n| 현장명 | 강남 |\n| 공종 | 전기 |\n\n"
        "## ■ 사전조사 결과\n\n| 항목 | 내용 |\n|------|------|\n| 확인사항 | 없음 |\n"
    ),
    "created_at": "2026-08-05 10:00:00",
}

# 2026-08-05 2차 피드백: "박스 안 텍스트가 셀 크기 때문에 잘려서 안 보이는 게
# 없도록" — 열너비 단위(통합문서 기본폰트 Calibri 11pt 기준)와 실제 본문
# 폰트(12pt, 한글은 맑은 고딕 등으로 자동 치환되어 더 넓음) 사이의 폭 추정
# 오차 때문에 병합셀 행높이가 실제보다 낮게 계산되는 경우를 검증한다.
# "내용" 열 폭(TBM 일지 스펙 2번째 값=22단위) 기준, 안전마진 적용 전에는
# 2줄로 계산되지만 적용 후에는 3줄로 계산되어야 하는 경계값(한글 19자=38단위).
SAMPLE_RECORD_ROW_HEIGHT_SAFETY = {
    "id": "rowheight_safety1",
    "document_type": "TBM 일지",
    "project_info": "행 높이 안전마진 검증용",
    "draft": "| 항목 | 내용 |\n|------|------|\n| 테스트 | " + "가" * 19 + " |\n",
    "created_at": "2026-08-05 10:00:00",
}

# 2026-08-05 3차 피드백: "(빈칸 - 현장 기재)" 같은 안내문은 실제 값이 아니므로
# 연한 회색으로 구분 표시 — 표 셀(전체가 플레이스홀더/일부만 플레이스홀더 둘 다),
# 헤딩, 서술형 텍스트 블록 모두에서 확인한다.
SAMPLE_RECORD_PLACEHOLDER = {
    "id": "placeholder1",
    "document_type": "TBM 일지",
    "project_info": "플레이스홀더 회색 처리 검증용",
    "draft": (
        "| 항목 | 내용 |\n|------|------|\n"
        "| 작업일자 | (빈칸 - 현장 기재) |\n"
        "| 작성자 | 김철수, (빈칸 - 현장 기재) |\n\n"
        "## ■ 비고\n\n"
        "특이사항 없음. 서명: (빈칸 - 현장 기재)\n"
    ),
    "created_at": "2026-08-05 10:00:00",
}

# 2026-08-05 요청: "표준 작업계획서 (전기 작업)"처럼 문서 제목 옆에 세부
# 작업유형을 표기 — work_type이 record에 있을 때만 붙는다.
SAMPLE_RECORD_WORKPLAN_WITH_WORKTYPE = {
    "id": "worktype1",
    "document_type": "표준 작업계획서",
    "project_info": "작업유형 표기 검증용",
    "work_type": "전기작업",
    "draft": "| 항목 | 내용 |\n|------|------|\n| 현장명 | 강남 |\n",
    "created_at": "2026-08-05 10:00:00",
}


def run():
    results = []

    xlsx_bytes = record_to_xlsx_bytes(SAMPLE_RECORD)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    results.append(("시트명이 문서종류로 설정됨", ws.title == "위험성평가표"))

    # --- 공통 1: 문서 제목(PDF와 동일하게 28pt·굵게·밑줄)이 맨 위에 렌더링됨 ---
    results.append(("문서 제목이 1행 B열에 렌더링됨", ws.cell(row=1, column=2).value == "위험성평가표"))
    results.append(("문서 제목 폰트 28pt·굵게·밑줄", (
        ws.cell(row=1, column=2).font.size == 28
        and ws.cell(row=1, column=2).font.bold is True
        and ws.cell(row=1, column=2).font.underline == "single"
    )))
    results.append(("문서 제목이 가운데정렬됨", ws.cell(row=1, column=2).alignment.horizontal == "center"))

    # --- 공통 3: A열은 공란, 모든 내용은 B열부터 ---
    results.append(("A열(1열)은 어느 행에도 값이 없음(공란)", not any(
        ws.cell(row=r, column=1).value not in (None, "") for r in range(1, ws.max_row + 1)
    )))

    # --- 박스 제목(헤딩)이 표 앞에 렌더링됨, 이제 B열부터 + 16pt(공통 2) ---
    results.append(("첫 박스 제목(헤딩)이 2행 B열에 렌더링됨", ws.cell(row=2, column=2).value == "■ 기본 정보"))
    results.append(("박스 제목 폰트 16pt·굵게·검정", (
        ws.cell(row=2, column=2).font.size == 16
        and ws.cell(row=2, column=2).font.bold is True
        and ws.cell(row=2, column=2).font.color.rgb.endswith("000000")
    )))
    results.append(("레벨1 제목('위험성평가표 초안')은 렌더링 안 됨(문서 제목과 중복)", not any(
        ws.cell(row=r, column=2).value == "위험성평가표 초안" for r in range(1, 15)
    )))

    results.append((
        "첫 번째 표 헤더 위치(헤딩 다음 행인 3행, B3/C3) 확인",
        ws.cell(row=3, column=2).value == "항목" and ws.cell(row=3, column=3).value == "내용",
    ))
    results.append((
        "굵게(**) 제거된 셀 값 확인(표 헤더 다음 행인 4행)",
        ws.cell(row=4, column=2).value == "현장명" and ws.cell(row=4, column=3).value == "강남지사_광케이블",
    ))
    results.append(("표 헤더 행(3행) 볼드+12pt 스타일 적용 확인", (
        ws.cell(row=3, column=2).font.bold is True and ws.cell(row=3, column=2).font.size == 12
    )))
    results.append(("표 데이터 셀(4행)도 12pt", ws.cell(row=4, column=3).font.size == 12))

    # --- 두 번째 박스는 첫 박스(헤딩1+표3행+빈행1=4행) 다음인 7행부터 ---
    results.append(("두 번째 박스 제목이 7행 B열에 렌더링됨", ws.cell(row=7, column=2).value == "■ 위험요인"))
    results.append((
        "두 번째 표 헤더 위치(헤딩 다음 행인 8행) 확인",
        ws.cell(row=8, column=2).value == "순번" and ws.cell(row=8, column=3).value == "위험요인",
    ))

    # --- kv표(항목/내용, 2열)와 다른 표(3열)가 섞여 있으면, kv표의 "내용" 칸을
    # 전체 표 중 가장 넓은 열 개수만큼 병합해서 박스 가로 크기를 맞춘다
    # (B열부터 시작하므로 병합 범위도 한 칸씩 밀림: C3:D3, C4:D4) ---
    merge_ranges = [str(r) for r in ws.merged_cells.ranges]
    results.append(("kv표 헤더 행의 '내용' 칸이 C3:D3로 병합되어 3열 표와 폭이 맞음", "C3:D3" in merge_ranges))
    results.append(("kv표 데이터 행(4행)의 값 칸도 C4:D4로 병합됨", "C4:D4" in merge_ranges))
    results.append(("박스 제목도 B열부터 전체 폭만큼 병합됨(B2:D2)", "B2:D2" in merge_ranges))

    # --- 표가 아예 없는(서술형 텍스트만 있는) 문서도 제목·B열 규칙을 그대로 따름 ---
    xlsx_bytes_empty = record_to_xlsx_bytes(SAMPLE_RECORD_NO_TABLE)
    wb2 = load_workbook(io.BytesIO(xlsx_bytes_empty))
    ws2 = wb2.active
    results.append(("표 없는 문서도 1행 B열에 문서 제목 렌더링", ws2.cell(row=1, column=2).value == "기타"))
    results.append((
        "표 없는 문서의 원문 텍스트는 제목 다음 행(2행) B열에 기록",
        ws2.cell(row=2, column=2).value == "이 문서에는 표가 없습니다.",
    ))

    # --- 표가 아닌 서술형 섹션("3. 중점(One Point) 위험요인" 등)도 사라지지
    # 않고 렌더링됨(2026-08-05 실사용 XLSX에서 통째로 빠졌던 버그) ---
    xlsx_bytes_prose = record_to_xlsx_bytes(SAMPLE_RECORD_WITH_HEADING_AND_PROSE)
    wb_prose = load_workbook(io.BytesIO(xlsx_bytes_prose))
    ws_prose = wb_prose.active
    all_values_prose = [
        ws_prose.cell(row=r, column=2).value
        for r in range(1, ws_prose.max_row + 1)
    ]
    results.append((
        "'3. 중점(One Point) 위험요인' 박스 제목이 렌더링됨",
        "3. 중점(One Point) 위험요인" in all_values_prose,
    ))
    results.append((
        "서술형 본문 내용('무전압 상태')이 실제로 렌더링됨",
        any(v is not None and "무전압 상태" in str(v) for v in all_values_prose),
    ))
    # 두 번째 박스(핵심 위험요인, 3열)가 있으므로 max_col_count=3 — 첫 박스의
    # kv표(항목/내용)도 C:D로 병합돼 폭이 맞아야 한다(B열 시작 기준).
    prose_merge_ranges = [str(r) for r in ws_prose.merged_cells.ranges]
    results.append((
        "박스 너비가 다른 표들과 안 맞던 kv표도 C:D 병합으로 폭이 맞춰짐",
        "C4:D4" in prose_merge_ranges,
    ))

    xlsx_bytes_score = record_to_xlsx_bytes(SAMPLE_RECORD_WITH_SCORE)
    wb3 = load_workbook(io.BytesIO(xlsx_bytes_score))
    ws3 = wb3.active

    # 이 표는 헤딩 없이 표 하나뿐 — 1행 제목, 2행 표헤더, 3행부터 데이터 (B열부터)
    results.append((
        "빈도·강도 AI 제안값이 순수 숫자로 저장됨(3행, C/D열)",
        ws3.cell(row=3, column=3).value == 3
        and ws3.cell(row=3, column=4).value == 2
        and isinstance(ws3.cell(row=3, column=3).value, int),
    ))
    results.append((
        "위험등급·개선후 위험등급 AI 제안값이 순수 등급 문자로 저장됨(E/F열)",
        ws3.cell(row=3, column=5).value == "A"
        and ws3.cell(row=3, column=6).value == "B",
    ))
    results.append((
        "AI 제안값 안내 문구는 셀 메모(comment)로 보존됨",
        ws3.cell(row=3, column=5).comment is not None
        and "AI 제안값" in ws3.cell(row=3, column=5).comment.text,
    ))
    results.append((
        "위험등급 열에 A/B/C 등급 조건부서식이 걸림",
        any(
            rule.formula == ['"A"']
            for rules in ws3.conditional_formatting._cf_rules.values()
            for rule in rules
        ),
    ))
    results.append((
        "점수가 아닌 일반 텍스트 셀은 그대로 문자열 유지(B열)",
        ws3.cell(row=3, column=2).value == "지게차 충돌"
        and ws3.cell(row=3, column=2).comment is None,
    ))
    # --- AI 제안값이 있는 표는 PDF/HWPX와 동일하게 각주 문구가 표 아래에
    # 추가됨(2026-08-05 요청 — 지금까지 XLSX만 이 각주가 빠져 있었음) ---
    score_all_values = [ws3.cell(row=r, column=2).value for r in range(1, ws3.max_row + 1)]
    results.append((
        "AI 제안값 각주(AI_SCORE_FOOTNOTE)가 표 아래에 렌더링됨",
        AI_SCORE_FOOTNOTE in score_all_values,
    ))

    # --- "위험성 추정 행렬" 참고표는 헤더명이 "위험등급"이 아니라
    # risk_grade_column_indices로는 안 잡히지만, 셀 값 자체가 등급 문자면
    # 본문 위험성평가표와 같은 색으로 칠해져야 한다(2026-08-05 요청) ---
    xlsx_bytes_matrix = record_to_xlsx_bytes(SAMPLE_RECORD_RISK_MATRIX)
    wb_matrix = load_workbook(io.BytesIO(xlsx_bytes_matrix))
    ws_matrix = wb_matrix.active
    risk_style = STYLE_SPECS["위험성평가표"]
    results.append((
        "위험성 추정 행렬의 'A' 셀(4행 E열)에 A등급 색이 직접 채워짐",
        ws_matrix.cell(row=4, column=5).fill.fgColor.rgb.upper().endswith(risk_style.risk_grade_colors["A"]),
    ))
    results.append((
        "위험성 추정 행렬의 'B' 셀(4행 D열)에 B등급 색이 직접 채워짐",
        ws_matrix.cell(row=4, column=4).fill.fgColor.rgb.upper().endswith(risk_style.risk_grade_colors["B"]),
    ))
    results.append((
        "행 라벨 열(B열, '1(낮음)' 등)은 A/B/C가 아니라 색칠되지 않음",
        ws_matrix.cell(row=3, column=2).fill.fgColor.rgb in (None, "00000000"),
    ))

    # --- 위험성평가표 전용: 열 폭을 셀 내용 글자수 기반으로 자동조절 ---
    xlsx_bytes_width = record_to_xlsx_bytes(SAMPLE_RECORD_RISK_WIDE_NARROW)
    wb4 = load_workbook(io.BytesIO(xlsx_bytes_width))
    ws4 = wb4.active
    # 표만 있고 헤딩 없음 → 1행 제목, 2행 표헤더 → 열: B=단위작업, C=빈도, D=위험성 감소대책
    narrow_width = ws4.column_dimensions["C"].width  # 빈도 (짧은 숫자)
    wide_width = ws4.column_dimensions["D"].width  # 위험성 감소대책 (긴 문장)
    results.append(("위험성평가표: 내용이 짧은 '빈도' 열보다 긴 '위험성 감소대책' 열이 훨씬 넓음", wide_width > narrow_width * 3))
    static_spec_widths = STYLE_SPECS["위험성평가표"].column_widths
    results.append((
        "위험성평가표 열폭은 정적 스펙 그대로가 아니라 실제 조정됨",
        [ws4.column_dimensions[c].width for c in ("B", "C", "D")] != static_spec_widths[:3],
    ))

    # --- 다른 문서유형(TBM 일지 등)은 여전히 정적 스펙을 그대로 씀(스코프 밖) ---
    xlsx_bytes_other = record_to_xlsx_bytes(SAMPLE_RECORD_OTHER_DOC_TYPE)
    wb5 = load_workbook(io.BytesIO(xlsx_bytes_other))
    ws5 = wb5.active
    tbm_spec_widths = STYLE_SPECS["TBM 일지"].column_widths
    results.append((
        "TBM 일지 등은 여전히 document_styles의 정적 열폭 스펙을 그대로 씀",
        ws5.column_dimensions["B"].width == tbm_spec_widths[0]
        and ws5.column_dimensions["C"].width == tbm_spec_widths[1],
    ))

    # --- 병합된 셀도 줄바꿈 내용에 맞춰 행 높이가 자동으로 늘어남
    # (2026-08-05 요청 — 안 그러면 긴 문단이 한 줄로 눌려 보이거나 "내용이
    # 빠졌다"고 오해하게 됨) ---
    xlsx_bytes_long = record_to_xlsx_bytes(SAMPLE_RECORD_LONG_TEXT)
    wb6 = load_workbook(io.BytesIO(xlsx_bytes_long))
    ws6 = wb6.active
    # 1행=제목, 2행=박스제목("종합의견"), 3행=긴 서술형 본문
    long_text_row_height = ws6.row_dimensions[3].height
    results.append(("긴 서술형 문단이 있는 행은 기본 높이(15pt)보다 훨씬 큼", (
        long_text_row_height is not None and long_text_row_height > 40
    )))
    # 6행=kv표 헤더, 7행="짧은값|예"(둘 다 짧음), 8행="긴값|매설물 관리기관..."(내용 열이 김)
    short_row_height = ws6.row_dimensions[7].height
    long_kv_row_height = ws6.row_dimensions[8].height
    results.append((
        "kv표에서도 내용이 긴 행이 짧은 행보다 높이가 훨씬 큼",
        long_kv_row_height is not None and short_row_height is not None
        and long_kv_row_height > short_row_height * 1.5,
    ))

    # --- 2026-08-05 요청: 문서 전체가 A4 용지에 맞게 출력되도록 용지 크기를
    # 명시하고, 표준 작업계획서·TBM 일지는 PDF와 동일하게 세로방향으로 낸다 ---
    results.append(("위험성평가표는 가로방향 유지", ws.page_setup.orientation == "landscape"))
    results.append(("위험성평가표도 A4 용지 크기가 명시됨", ws.page_setup.paperSize == 9))
    results.append(("TBM 일지는 세로방향(A4 세로) 적용", ws_prose.page_setup.orientation == "portrait"))
    results.append(("TBM 일지도 A4 용지 크기가 명시됨", ws_prose.page_setup.paperSize == 9))

    # --- "엑셀 기능중에 틀고정 기능은 빼줘"(위험성평가표) — 표가 많아 스크롤이
    # 잦은 문서라 틀고정이 오히려 헤더를 가려서 혼란을 준다는 피드백 ---
    xlsx_bytes_freeze_risk = record_to_xlsx_bytes(SAMPLE_RECORD_FREEZE_RISK)
    ws_freeze_risk = load_workbook(io.BytesIO(xlsx_bytes_freeze_risk)).active
    results.append(("위험성평가표는 틀고정을 쓰지 않음", ws_freeze_risk.freeze_panes is None))
    results.append(("SAMPLE_RECORD(위험성평가표)도 틀고정 없음", ws.freeze_panes is None))

    # --- "작업 개요 박스 다음에 틀고정 기능을 넣어줘"(표준 작업계획서) — 문서가
    # 길어서 스크롤하면 작업 개요가 안 보인다는 피드백. SAMPLE_RECORD_FREEZE_WORKPLAN은
    # 1행 제목, 2행 헤딩("■ 작업 개요"), 3행 표헤더, 4~5행 데이터, 6행 빈행(표 사이
    # 간격) 다음인 7행부터가 두 번째 박스 — 그 경계인 7행 위쪽을 고정한다.
    xlsx_bytes_freeze_wp = record_to_xlsx_bytes(SAMPLE_RECORD_FREEZE_WORKPLAN)
    ws_freeze_wp = load_workbook(io.BytesIO(xlsx_bytes_freeze_wp)).active
    results.append(("표준 작업계획서는 첫 번째 박스(작업 개요) 바로 다음 행에 틀고정", ws_freeze_wp.freeze_panes == "A7"))
    results.append(("표준 작업계획서는 세로방향(A4 세로) 적용", ws_freeze_wp.page_setup.orientation == "portrait"))
    results.append(("표준 작업계획서도 A4 용지 크기가 명시됨", ws_freeze_wp.page_setup.paperSize == 9))

    # --- TBM 일지는 세로방향으로 바뀌어도 틀고정 위치는 기존 방식(두 번째 표
    # 헤더 다음) 그대로 유지되어야 한다(작업계획서만 별도 요청된 사항) ---
    results.append(("TBM 일지는 틀고정 위치가 기존 방식(두 번째 표 헤더 다음) 그대로 유지됨", ws_prose.freeze_panes == "A8"))

    # --- "표준 작업계획서 제목 옆에 세부 작업명을 표기해줘. 예) 표준 작업계획서
    # (전기 작업)" — record에 work_type이 있으면 제목에 괄호로 덧붙인다 ---
    xlsx_bytes_wt = record_to_xlsx_bytes(SAMPLE_RECORD_WORKPLAN_WITH_WORKTYPE)
    ws_wt = load_workbook(io.BytesIO(xlsx_bytes_wt)).active
    results.append(("work_type이 있으면 문서 제목에 괄호로 덧붙음", ws_wt.cell(row=1, column=2).value == "표준 작업계획서 (전기작업)"))
    results.append(("work_type이 없는 기존 표준 작업계획서는 제목이 그대로 유지됨", ws_freeze_wp.cell(row=1, column=2).value == "표준 작업계획서"))

    # --- 2026-08-05 2차 피드백 1: 셀 크기 때문에 텍스트가 잘리는 문제 —
    # 열너비 단위 추정과 실제 렌더링 폭 사이의 오차에 안전마진을 적용해
    # 행 높이를 더 넉넉하게 계산해야 한다. "내용" 열(TBM 스펙 22단위)에
    # 한글 19자(38단위)를 넣으면, 안전마진 적용 전엔 2줄(ceil(38/22)=2)로
    # 계산되지만 적용 후엔 3줄(ceil(38*1.25/22)=3)로 계산되어야 한다 —
    # 2줄 높이(구 상수 기준 2*15+5=35)보다 확실히 큰 45pt를 기준으로 확인.
    xlsx_bytes_rowsafety = record_to_xlsx_bytes(SAMPLE_RECORD_ROW_HEIGHT_SAFETY)
    ws_rowsafety = load_workbook(io.BytesIO(xlsx_bytes_rowsafety)).active
    # 1행=제목, 2행=표헤더, 3행=데이터("테스트"/한글19자)
    results.append((
        "안전마진 적용으로 행 높이가 2줄 계산치보다 넉넉하게(3줄 이상) 잡힘",
        ws_rowsafety.row_dimensions[3].height is not None and ws_rowsafety.row_dimensions[3].height > 45,
    ))

    # --- 2026-08-05 2차 피드백 2: 인쇄 여백을 상하좌우 25px(≈0.26인치)로 축소 ---
    for name_label, target_ws in (("위험성평가표", ws), ("TBM 일지", ws_prose), ("표준 작업계획서", ws_freeze_wp)):
        margin_ok = all(
            abs(getattr(target_ws.page_margins, side) - 25 / 96) < 0.01
            for side in ("left", "right", "top", "bottom")
        )
        results.append((f"{name_label}: 인쇄 여백이 상하좌우 25px(≈0.26인치)로 축소됨", margin_ok))

    # --- 2026-08-05 3차 피드백: fitToWidth 자동 맞춤을 뷰어가 지키지 않아
    # 인쇄 미리보기에서 내용이 잘리는 게 실제로 확인됨 — 이제는 항상 명시적
    # 배율을 계산해서 적용하고(fitToPage는 항상 꺼짐), 그 배율은
    # _print_scale_percent가 실제로 계산한 값과 정확히 일치해야 한다
    # (안전마진을 반영해 100%를 넘겨 확대하지는 않는다 — 2차 수정 때 TBM
    # 일지를 124%로 확대했다가 실제 인쇄에서 오히려 넘쳐 잘렸던 회귀 방지).
    # SAMPLE_RECORD_WITH_HEADING_AND_PROSE(ws_prose)의 표 중 가장 열이 많은
    # 것은 3열("핵심 위험요인" 표) — 실제 record_to_xlsx_bytes 내부에서 쓰는
    # max_col_count와 동일하게 맞춰야 계산값이 일치한다.
    tbm_widths = STYLE_SPECS["TBM 일지"].column_widths
    expected_tbm_scale = _print_scale_percent("TBM 일지", tbm_widths, 3)
    results.append(("TBM 일지는 배율을 직접 지정하므로 fitToPage(자동 폭맞춤)는 꺼짐", ws_prose.sheet_properties.pageSetUpPr.fitToPage is not True))
    results.append(("TBM 일지에 적용된 배율이 _print_scale_percent 계산값과 일치", ws_prose.page_setup.scale == expected_tbm_scale))

    # 위험성평가표(실제 13열 스펙)도 이제 fitToWidth가 아니라 명시적 배율을 쓴다 —
    # 열이 많아 폭이 넉넉하므로 배율은 100% 이하로 계산되어야 한다.
    risk_full_widths = STYLE_SPECS["위험성평가표"].column_widths
    risk_full_scale = _print_scale_percent("위험성평가표", risk_full_widths, len(risk_full_widths))
    results.append(("위험성평가표(실제 13열 스펙)는 배율이 100% 이하로 계산됨", risk_full_scale <= 100))
    results.append(("계산된 배율은 최소/최대 한도 안에 있음(50~115)", 50 <= risk_full_scale <= 115 and 50 <= expected_tbm_scale <= 115))

    # --- 2026-08-05 3차 피드백: "(빈칸 - 현장 기재)" 안내문 회색 처리 ---
    xlsx_bytes_ph = record_to_xlsx_bytes(SAMPLE_RECORD_PLACEHOLDER)
    ws_ph_plain = load_workbook(io.BytesIO(xlsx_bytes_ph)).active
    ws_ph_rich = load_workbook(io.BytesIO(xlsx_bytes_ph), rich_text=True).active
    # 1행=제목, 2행=표헤더, 3행="작업일자|(빈칸...)", 4행="작성자|김철수, (빈칸...)",
    # 5행=빈행, 6행=헤딩("■ 비고"), 7행=서술형 텍스트("...서명: (빈칸...)")
    results.append((
        "전체가 플레이스홀더인 셀도 원문 텍스트 값 자체는 그대로 보존됨",
        ws_ph_plain.cell(row=3, column=3).value == "(빈칸 - 현장 기재)",
    ))
    results.append((
        "전체가 플레이스홀더인 셀은 rich text로 회색 처리됨",
        _placeholder_run_present(ws_ph_rich.cell(row=3, column=3).value),
    ))
    embedded_rich = ws_ph_rich.cell(row=4, column=3).value
    results.append(("일부만 플레이스홀더인 셀도 그 부분만 회색 처리됨", _placeholder_run_present(embedded_rich)))
    results.append((
        "일부만 플레이스홀더인 셀은 앞부분(실제 값)은 일반 색 그대로 유지됨",
        isinstance(embedded_rich, CellRichText) and any(
            "김철수" in b.text and not (getattr(getattr(b.font, "color", None), "rgb", "") or "").upper().endswith("999999")
            for b in embedded_rich
        ),
    ))
    results.append((
        "플레이스홀더 없는 박스 제목은 평범한 문자열 그대로 유지됨(불필요한 rich text 안 씀)",
        ws_ph_plain.cell(row=6, column=2).value == "■ 비고"
        and not isinstance(ws_ph_plain.cell(row=6, column=2).value, CellRichText),
    ))
    results.append((
        "서술형 텍스트 블록 안의 플레이스홀더도 회색 처리됨",
        _placeholder_run_present(ws_ph_rich.cell(row=7, column=2).value),
    ))

    all_ok = True
    for name, ok in results:
        print(f"[{'PASS' if ok else 'FAIL'}] {name}")
        all_ok = all_ok and ok

    print()
    print("전체 결과:", "PASS" if all_ok else "FAIL (위 로그 확인)")
    return all_ok


if __name__ == "__main__":
    run()
